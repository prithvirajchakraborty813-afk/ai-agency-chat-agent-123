#!/usr/bin/env python3
"""
proposal_generator.py — Agent #5 in the 14-agent plan: the Proposal
Generator. Sits right after gemini_vertex_qualifier.py in the pipeline.

WHAT THIS DOES: takes qualified.csv (output of gemini_vertex_qualifier.py)
and, for every row marked qualified=True, asks Gemini to (1) recommend
which of your three pricing tiers (starter/growth/custom, defined in
tiers_agency.json) best fits that specific business, and (2) draft a
short, specific, WhatsApp/email-ready proposal message referencing their
actual pain point and the recommended tier's real price.

IMPORTANT DESIGN CHOICE: the model is given the exact tier names and
prices from tiers_agency.json and told to use them verbatim — it does
NOT invent pricing. This avoids the same class of problem the domain
field had (the model filling in text you didn't ask for): a proposal
with a hallucinated price is worse than a proposal with none.

    gemini_vertex_qualifier.py --> qualified.csv (qualified=True rows)
                                          |
                                          v
    proposal_generator.py       --> proposals.csv (tier + proposal text)

Setup: identical to gemini_vertex_qualifier.py —
    1. gcloud auth application-default login
    2. gcloud config set project YOUR_PROJECT_ID
    3. pip install requests google-auth

Usage:
    python proposal_generator.py --in qualified.csv --out proposals.csv \
        --project project-95bf86c6-f889-4996-ba3
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from dataclasses import dataclass

import requests
import google.auth
import google.auth.transport.requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

LOCATION = "us-central1"
VERTEX_ENDPOINT_TEMPLATE = (
    "https://{location}-aiplatform.googleapis.com/v1/projects/{project}"
    "/locations/{location}/publishers/google/models/{model}:generateContent"
)
DEFAULT_MODEL = "gemini-2.5-flash-lite"  # same confirmed-live model as the rest of the pipeline
AUTH_SCOPES = ["https://www.googleapis.com/auth/cloud-platform"]


@dataclass
class Proposal:
    name: str
    domain: str
    phone: str
    recommended_tier: str
    setup_price: str
    monthly_price: str
    reasoning: str
    proposal_text: str


RESPONSE_SCHEMA = {
    "type": "OBJECT",
    "properties": {
        "recommended_tier": {
            "type": "STRING",
            "description": "Must be exactly one of the tier keys given in the prompt (e.g. 'starter', 'growth', 'custom') — do not invent a new tier name.",
        },
        "reasoning": {"type": "STRING", "description": "1-2 sentences on why this tier fits this business."},
        "proposal_text": {
            "type": "STRING",
            "description": "The proposal message, ready to send as-is over WhatsApp or email.",
        },
    },
    "required": ["recommended_tier", "reasoning", "proposal_text"],
}

PROPOSAL_SYSTEM_PROMPT = """You are a proposal-writing assistant for a solo AI automation agency. \
The agency builds custom AI agents (WhatsApp or website bots) for small Indian businesses — \
the bot answers customer FAQs, captures booking/inquiry details, and sends follow-ups.

You will be given: a business's name, domain, industry/pain-point notes, and a list of pricing \
tiers with their exact prices and descriptions.

Your job:
1. Recommend the ONE tier that best fits this specific business's size and complexity. Use the \
   tier key exactly as given (e.g. "starter"), never invent a new one.
2. Write a short proposal message, ready to send as-is over WhatsApp or email. It must:
   - Reference something concrete and specific about this business (not generic filler)
   - Briefly describe what would actually be built for them (tie it to their stated pain point)
   - State the recommended tier's price EXACTLY as given to you — never round, estimate, or invent a number
   - Invite them to a short async reply or call to confirm details
   - Read as if written by a real person doing genuine research on their business — it must not \
     falsely claim no automation was involved anywhere in this process, but it also shouldn't read \
     like a mass-blasted template."""


def load_tiers(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def tiers_prompt_block(tiers: dict) -> str:
    lines = []
    for key, t in tiers.items():
        lines.append(
            f"- \"{key}\" ({t['label']}): {t['setup_price']}, {t['monthly_price']}. {t['description']}"
        )
    return "\n".join(lines)


class VertexGeminiClient:
    def __init__(self, project: str, model: str = DEFAULT_MODEL, location: str = LOCATION):
        self.project = project
        self.model = model
        self.location = location
        self.endpoint = VERTEX_ENDPOINT_TEMPLATE.format(location=location, project=project, model=model)
        try:
            self.credentials, _ = google.auth.default(scopes=AUTH_SCOPES)
        except Exception as e:
            raise ValueError(
                "Could not find Application Default Credentials. Run "
                "'gcloud auth application-default login' first, then retry.\n"
                f"(underlying error: {e})"
            )

    def _access_token(self) -> str:
        request = google.auth.transport.requests.Request()
        self.credentials.refresh(request)
        return self.credentials.token

    def generate_json(self, system: str, user: str, schema: dict,
                       temperature: float = 0.4, max_retries: int = 6) -> dict:
        payload = {
            "systemInstruction": {"parts": [{"text": system}]},
            "contents": [{"role": "user", "parts": [{"text": user}]}],
            "generationConfig": {
                "temperature": temperature,
                "maxOutputTokens": 800,
                "responseMimeType": "application/json",
                "responseSchema": schema,
            },
        }
        last_err = None
        for attempt in range(max_retries):
            try:
                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {self._access_token()}",
                }
                resp = requests.post(self.endpoint, headers=headers, json=payload, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                text = data["candidates"][0]["content"]["parts"][0]["text"]
                return json.loads(text)
            except requests.exceptions.HTTPError as e:
                last_err = e
                status = e.response.status_code if e.response is not None else None
                if status == 429:
                    # Rate limit (quota-per-minute), not a real failure — the old
                    # 1s/2s/4s backoff wasn't long enough to clear the window, so
                    # a 429 would burn all 3 attempts and the row got dropped
                    # (this is what happened to Shamz Clinic). Respect a
                    # Retry-After header if the API sends one; otherwise back off
                    # much more aggressively (starts at 10s, caps at 90s).
                    retry_after = None
                    if e.response is not None:
                        retry_after = e.response.headers.get("Retry-After")
                    if retry_after:
                        wait = float(retry_after)
                    else:
                        wait = min(10 * (2 ** attempt), 90)
                    print(f"  [warn] Rate limited (429); waiting {wait:.0f}s before retry "
                          f"{attempt + 1}/{max_retries}...", file=sys.stderr)
                else:
                    wait = min(2 ** attempt, 30)
                    print(f"  [warn] Vertex AI call failed ({e}); retrying in {wait}s...", file=sys.stderr)
                time.sleep(wait)
            except Exception as e:
                last_err = e
                wait = min(2 ** attempt, 30)
                print(f"  [warn] Vertex AI call failed ({e}); retrying in {wait}s...", file=sys.stderr)
                time.sleep(wait)
        raise RuntimeError(f"Vertex AI call failed after {max_retries} attempts: {last_err}")


def generate_proposal(client: VertexGeminiClient, row: dict, tiers: dict) -> Proposal:
    user_prompt = f"""Business to propose to:
  Name: {row.get('name', '')}
  Domain: {row.get('domain', '')}
  Likely pain point: {row.get('pain_points_guess', '')}
  Qualifier's notes: {row.get('reasoning', '')}

Available tiers:
{tiers_prompt_block(tiers)}
"""
    parsed = client.generate_json(PROPOSAL_SYSTEM_PROMPT, user_prompt, RESPONSE_SCHEMA)

    tier_key = str(parsed.get("recommended_tier", "")).strip().lower()
    tier_info = tiers.get(tier_key)
    if tier_info is None:
        # Model returned a tier key that isn't in tiers_agency.json — don't
        # silently trust a price that doesn't exist. Fall back to "starter"
        # (the safest, cheapest default) and flag it loudly instead of
        # writing a proposal with a made-up price.
        print(f"  [warn] Model returned unknown tier '{tier_key}' for {row.get('name', '?')} "
              f"— falling back to 'starter'.", file=sys.stderr)
        tier_key = "starter"
        tier_info = tiers["starter"]

    return Proposal(
        name=row.get("name", ""),
        domain=row.get("domain", ""),
        phone=row.get("phone", ""),
        recommended_tier=tier_info["label"],
        setup_price=tier_info["setup_price"],
        monthly_price=tier_info["monthly_price"],
        reasoning=parsed.get("reasoning", ""),
        proposal_text=parsed.get("proposal_text", ""),
    )


def write_xlsx(proposals: list[Proposal], out_xlsx: str) -> None:
    """Same data as the CSV, but actually readable when opened in Excel:
    fixed column widths per field (narrow for name/phone, wide for the
    proposal text), wrapped text so rows aren't one giant unreadable line,
    a bold frozen header, and light banding so rows are easy to track
    across the wide columns."""
    headers = ["Business Name", "Domain", "Phone", "Recommended Tier", "Setup Price",
               "Monthly Price", "Why This Tier", "Proposal Text (ready to send)"]
    widths = [26, 20, 16, 16, 22, 20, 45, 70]

    wb = Workbook()
    ws = wb.active
    ws.title = "Proposals"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    band_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for i, p in enumerate(proposals, start=2):
        values = [p.name, p.domain, p.phone, p.recommended_tier, p.setup_price,
                  p.monthly_price, p.reasoning, p.proposal_text]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = body_font
            cell.alignment = wrap
            if i % 2 == 0:
                cell.fill = band_fill
        # Give each row enough height for the longest wrapped field (the
        # proposal text column), roughly 15 chars per line at this width.
        longest = max(len(str(v)) for v in values)
        ws.row_dimensions[i].height = max(30, min(200, (longest // 70 + 1) * 15))

    wb.save(out_xlsx)


def run(in_csv: str, out_csv: str, tiers_path: str, project: str,
        model: str = DEFAULT_MODEL, delay: float = 1.5) -> list[Proposal]:
    tiers = load_tiers(tiers_path)
    client = VertexGeminiClient(project=project, model=model)

    with open(in_csv, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    qualified_rows = [r for r in rows if str(r.get("qualified", "")).strip().lower() == "true"]
    skipped = len(rows) - len(qualified_rows)
    if skipped:
        print(f"Skipping {skipped} row(s) not marked qualified=True.")

    print(f"Generating proposals for {len(qualified_rows)} qualified lead(s) via Vertex AI Gemini ({model})...")
    proposals: list[Proposal] = []
    failed_rows: list[dict] = []
    for i, row in enumerate(qualified_rows, 1):
        try:
            p = generate_proposal(client, row, tiers)
            proposals.append(p)
            print(f"  [{i}/{len(qualified_rows)}] {p.name:30s} -> {p.recommended_tier}")
        except Exception as e:
            print(f"  [{i}/{len(qualified_rows)}] {row.get('name', '?')} FAILED: {e}", file=sys.stderr)
            failed_rows.append(row)
        time.sleep(delay)

    # Second pass: rows that failed above most likely hit the rate limit near
    # the end of their own retry budget, not because anything is wrong with
    # them specifically. Give the quota window time to clear, then retry each
    # failed row once more before giving up on it for good.
    if failed_rows:
        cooldown = 20
        print(f"\n{len(failed_rows)} row(s) failed on the first pass — waiting {cooldown}s "
              f"then retrying them once more...", file=sys.stderr)
        time.sleep(cooldown)
        still_failed = []
        for i, row in enumerate(failed_rows, 1):
            try:
                p = generate_proposal(client, row, tiers)
                proposals.append(p)
                print(f"  [retry {i}/{len(failed_rows)}] {p.name:30s} -> {p.recommended_tier}")
            except Exception as e:
                print(f"  [retry {i}/{len(failed_rows)}] {row.get('name', '?')} FAILED AGAIN: {e}",
                      file=sys.stderr)
                still_failed.append(row.get("name", "?"))
            time.sleep(delay)
        if still_failed:
            print(f"\n[warn] Still failed after retry pass, not in output: {', '.join(still_failed)}",
                  file=sys.stderr)

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["name", "domain", "phone", "recommended_tier", "setup_price",
                          "monthly_price", "reasoning", "proposal_text"])
        for p in proposals:
            writer.writerow([p.name, p.domain, p.phone, p.recommended_tier, p.setup_price,
                              p.monthly_price, p.reasoning, p.proposal_text])

    out_xlsx = out_csv.rsplit(".", 1)[0] + ".xlsx"
    write_xlsx(proposals, out_xlsx)
    print(f"Also wrote a formatted spreadsheet to {out_xlsx} (easier to read than the CSV in Excel).")

    return proposals


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate tiered proposals for qualified leads via Gemini on Vertex AI.")
    parser.add_argument("--in", dest="in_csv", required=True, help="CSV from gemini_vertex_qualifier.py")
    parser.add_argument("--out", dest="out_csv", default="proposals.csv")
    parser.add_argument("--tiers", default="tiers_agency.json", help="Path to the pricing tiers config")
    parser.add_argument("--project", required=True, help="Google Cloud project ID")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="Gemini model id")
    parser.add_argument("--delay", type=float, default=1.5,
                         help="Seconds to wait between requests (raise this if you keep hitting 429s)")
    args = parser.parse_args()

    proposals = run(args.in_csv, args.out_csv, args.tiers, args.project, args.model, args.delay)
    print(f"\nDone. Wrote {len(proposals)} proposal(s) to {args.out_csv}")


if __name__ == "__main__":
    main()
